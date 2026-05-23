#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════╗
║          School ERP Data Formatter  v1.0.0                       ║
║  Modern school data cleaning, ERP formatting & class-wise        ║
║  export desktop software built with Python + CustomTkinter        ║
╚══════════════════════════════════════════════════════════════════╝
"""

# ── Standard library ─────────────────────────────────────────────────────────
import os
import re
import sys
import json
import queue
import threading
import subprocess
import copy
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List, Tuple, Any

# ── Third-party ───────────────────────────────────────────────────────────────
import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

APP_NAME    = "School ERP Data Formatter"
APP_VERSION = "1.0.0"
WIN_SIZE    = "1400x900"

COLUMN_ALIASES: Dict[str, List[str]] = {
    "Regd_no": [
        "regd no", "regd_no", "registration no", "registration number",
        "admission no", "admission number", "roll no", "roll number",
    ],
    "FullName": [
        "fullname", "full_name", "student name", "name",
        "studentname", "student_name", "full name",
    ],
    "Gender": ["gender", "sex"],
    "Father Name": [
        "father name", "father_name", "father",
        "father's name", "fathername", "fathers name",
    ],
    "Mother Name": [
        "mother name", "mother_name", "mother",
        "mother's name", "mothername", "mothers name",
    ],
    "CurrentClass": [
        "currentclass", "current class", "grade", "class",
        "current_class", "std", "standard", "class name",
    ],
    "Permanent Address": [
        "permanent address", "address", "permanentaddress",
        "permanent_address", "addr",
    ],
    "DOB": [
        "dob", "date of birth", "birthdate", "birth_date",
        "dateofbirth", "date_of_birth", "birth date",
    ],
    "Guardian Contact Number": [
        "guardian contact number", "contact number", "phone", "mobile",
        "phone_number", "phonenumber", "contactnumber", "guardian_contact",
        "contact", "mobile_number", "mobilenumber", "contact no", "phone no",
    ],
}

# Fields used to drive the ERP mapping UI (ERP-required fields)
ERP_MAPPING_FIELDS: List[str] = ["Regd_no"] + [
    field for field in COLUMN_ALIASES.keys() if field != "Regd_no"
] + ["Class", "Section"]
ERP_HEADERS: List[str] = [
    "Regd_no", "Name", "Email", "Gender",
    "Father's_Name", "Mother's_Name", "Father's_Phone",
    "Date_of_birth", "Address", "Academic_Year", "Status",
    "Guardian_Name", "Relation", "Guardian_email",
    "Address2", "Phone_Number", "roll_number", "blood_group", "shift",
]

DATE_FORMAT_OPTIONS: List[str] = [
    "MM/DD/YYYY", "MM-DD-YYYY", "MM.DD.YYYY", "MM DD YYYY", "MMDDYYYY",
    "DD/MM/YYYY", "DD-MM-YYYY", "DD.MM.YYYY", "DD MM YYYY", "DDMMYYYY",
    "YYYY/MM/DD", "YYYY-MM-DD", "YYYY.MM.DD", "YYYY MM DD", "YYYYMMDD",
    "YYYY/DD/MM", "YYYY-DD-MM", "YYYY.DD.MM", "YYYY DD MM", "YYYYDDMM",
    "MM/YYYY/DD", "MM-YYYY-DD", "MM.YYYY.DD", "MM YYYY DD", "MMYYYYDD",
    "DD/YYYY/MM", "DD-YYYY-MM", "DD.YYYY.MM", "DD YYYY MM", "DDYYYYMM",
    "M/D/YYYY", "M-D-YYYY", "M.D.YYYY", "M D YYYY", "MDYYYY",
    "D/M/YYYY", "D-M-YYYY", "D.M.YYYY", "D M YYYY", "DMYYYY",
    "YYYY/M/D", "YYYY-M-D", "YYYY.M.D", "YYYY M D", "YYYYMD",
    "YY/MM/DD", "YY-MM-DD", "YY.MM.DD", "YY MM DD", "YYMMDD",
]


def _split_date_format_label(label: str) -> List[str]:
    return re.findall(r"Y+|M+|D+|[^YMD]+", label)


def _date_label_to_strptime(label: str) -> str:
    parts: List[str] = []
    for chunk in _split_date_format_label(label):
        if re.fullmatch(r"Y+", chunk):
            if len(chunk) == 2:
                parts.append("%y")
            else:
                parts.append("%Y")
        elif re.fullmatch(r"M+", chunk):
            parts.append("%m")
        elif re.fullmatch(r"D+", chunk):
            parts.append("%d")
        else:
            parts.append(chunk)
    return "".join(parts)


def _format_datetime_by_label(dt: datetime, label: str) -> str:
    parts: List[str] = []
    for chunk in _split_date_format_label(label):
        if re.fullmatch(r"Y+", chunk):
            if len(chunk) == 1:
                parts.append(str(dt.year % 10))
            elif len(chunk) == 2:
                parts.append(f"{dt.year % 100:02d}")
            else:
                parts.append(f"{dt.year:0{len(chunk)}d}" if len(chunk) > 4 else f"{dt.year:04d}")
        elif re.fullmatch(r"M+", chunk):
            parts.append(str(dt.month) if len(chunk) == 1 else f"{dt.month:02d}")
        elif re.fullmatch(r"D+", chunk):
            parts.append(str(dt.day) if len(chunk) == 1 else f"{dt.day:02d}")
        else:
            parts.append(chunk)
    return "".join(parts)


def _parse_date_by_label(value: str, label: Optional[str]) -> Optional[datetime]:
    if not value or str(value).lower() in ("nan", "none", ""):
        return None

    text = str(value).strip()
    if not label or label == "Auto":
        for dayfirst in (True, False):
            parsed = pd.to_datetime(text, errors="coerce", dayfirst=dayfirst)
            if pd.notnull(parsed):
                return parsed.to_pydatetime()
        return None

    pattern = _date_label_to_strptime(label)
    try:
        return datetime.strptime(text, pattern)
    except Exception:
        return None

STEPS: List[Tuple[int, str]] = [
    (1, "📁  Import File"),
    (2, "🔍  Column Detection"),
    (3, "🧹  Data Cleaning"),
    (4, "⚙️   Settings"),
    (5, "🔄  ERP Format"),
    (6, "📚  Class Separation"),
    (7, "💾  Export"),
]

PRESET_DIR = Path.home() / ".school_erp_formatter"
PRESET_FILE = PRESET_DIR / "presets.json"
STATE_FILE  = PRESET_DIR / "autosave.json"
RECENT_FILE = PRESET_DIR / "recent.json"


class AppState:
    def __init__(self):
        self.raw_df:     Optional[pd.DataFrame] = None
        self.cleaned_df: Optional[pd.DataFrame] = None
        self.erp_df:     Optional[pd.DataFrame] = None
        self.class_groups: Dict[str, pd.DataFrame] = {}
        self.source_file: str = ""
        self.file_size:   str = ""
        self.column_mapping: Dict[str, str] = {}
        self.school_name:    str = ""
        self.school_domain:  str = "school.com"
        self.output_format:  str = "XLSX"
        self.regd_no_mode:   str = "=ROW()-1"
        self.regd_no_custom: str = ""
        self.regd_no_column: str = ""
        self.dob_column_format: Dict[str, str] = {}
        self.dob_desired_format: Dict[str, str] = {}
        self.academic_year:  str = "=1"
        self.output_folder:  str = ""
        self.import_stats:   Dict[str, Any] = {}
        self.cleaning_stats: Dict[str, Any] = {}
        self.address_keep_spaces: bool = True
        # undo removed

    # Undo feature removed per user request

    def save_preset(self, name: str):
        PRESET_DIR.mkdir(exist_ok=True)
        presets = {}
        if PRESET_FILE.exists():
            try:
                presets = json.loads(PRESET_FILE.read_text())
            except Exception:
                pass
        # maintain FIFO queue of presets (max 7). New or updated preset moves to the end.
        MAX_PRESETS = 7
        if name in presets:
            # update ordering by removing existing entry first
            presets.pop(name, None)
        elif len(presets) >= MAX_PRESETS:
            # remove oldest (first) entry
            try:
                oldest = next(iter(presets))
                presets.pop(oldest, None)
            except Exception:
                pass

        presets[name] = {
            "school_name":   self.school_name,
            "school_domain": self.school_domain,
            "output_format": self.output_format,
            "regd_no_mode":  self.regd_no_mode,
            "academic_year": self.academic_year,
            "address_keep_spaces": self.address_keep_spaces,
        }
        PRESET_FILE.write_text(json.dumps(presets, indent=2))

    def delete_preset(self, name: str) -> bool:
        if not PRESET_FILE.exists():
            return False
        try:
            presets = json.loads(PRESET_FILE.read_text())
            if name in presets:
                presets.pop(name, None)
                PRESET_FILE.write_text(json.dumps(presets, indent=2))
                return True
        except Exception:
            pass
        return False

    def rename_preset(self, old: str, new: str) -> bool:
        if not PRESET_FILE.exists():
            return False
        try:
            presets = json.loads(PRESET_FILE.read_text())
            if old not in presets or new in presets:
                return False
            presets[new] = presets.pop(old)
            PRESET_FILE.write_text(json.dumps(presets, indent=2))
            return True
        except Exception:
            return False

    def load_preset(self, name: str) -> bool:
        if not PRESET_FILE.exists():
            return False
        try:
            presets = json.loads(PRESET_FILE.read_text())
            if name not in presets:
                return False
            p = presets[name]
            self.school_name   = p.get("school_name", "")
            self.school_domain = p.get("school_domain", "school.com")
            self.output_format = p.get("output_format", "XLSX")
            self.regd_no_mode  = p.get("regd_no_mode", "=ROW()-1")
            self.academic_year = p.get("academic_year", "=1")
            self.address_keep_spaces = p.get("address_keep_spaces", True)
            return True
        except Exception:
            return False

    def list_presets(self) -> List[str]:
        if not PRESET_FILE.exists():
            return []
        try:
            return list(json.loads(PRESET_FILE.read_text()).keys())
        except Exception:
            return []

    @staticmethod
    def add_recent(path: str):
        PRESET_DIR.mkdir(exist_ok=True)
        recent: List[str] = []
        if RECENT_FILE.exists():
            try:
                recent = json.loads(RECENT_FILE.read_text())
            except Exception:
                pass
        if path in recent:
            recent.remove(path)
        recent.insert(0, path)
        RECENT_FILE.write_text(json.dumps(recent[:10]))

    @staticmethod
    def get_recent() -> List[str]:
        if not RECENT_FILE.exists():
            return []
        try:
            return [p for p in json.loads(RECENT_FILE.read_text()) if os.path.exists(p)]
        except Exception:
            return []


class Logger:
    LEVELS = ("INFO", "OK", "WARN", "ERROR")

    def __init__(self, ui_callback):
        self._cb = ui_callback

    def _emit(self, msg: str, level: str):
        entry = {
            "ts":    datetime.now().strftime("%H:%M:%S"),
            "msg":   msg,
            "level": level,
        }
        try:
            self._cb(entry)
        except Exception:
            pass

    def info(self, msg: str):    self._emit(msg, "INFO")
    def success(self, msg: str): self._emit(msg, "OK")
    def warn(self, msg: str):    self._emit(msg, "WARN")
    def error(self, msg: str):   self._emit(msg, "ERROR")


class ImportEngine:
    @staticmethod
    def load(path: str, logger: Logger) -> Tuple[pd.DataFrame, dict]:
        logger.info(f"Loading: {os.path.basename(path)}")
        ext    = Path(path).suffix.lower()
        fsize  = os.path.getsize(path)
        size_s = (f"{fsize/1024:.1f} KB" if fsize < 1_048_576
                  else f"{fsize/1_048_576:.2f} MB")

        if ext in (".xlsx", ".xls"):
            xl     = pd.ExcelFile(path)
            sheets = xl.sheet_names
            logger.info(f"Sheets detected: {', '.join(sheets)}")
            df = pd.read_excel(path, sheet_name=sheets[0], dtype=str)
            for sheet in sheets[1:]:
                candidate = pd.read_excel(path, sheet_name=sheet, dtype=str)
                if len(candidate) > len(df):
                    df = candidate
                    logger.info(f"Using larger sheet: '{sheet}'")
                    break
        elif ext == ".csv":
            sheets = ["CSV"]
            for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
                try:
                    df = pd.read_csv(path, dtype=str, encoding=enc)
                    logger.info(f"CSV encoding: {enc}")
                    break
                except UnicodeDecodeError:
                    continue
            else:
                df = pd.read_csv(path, dtype=str, encoding="latin-1", errors="replace")
        else:
            raise ValueError(f"Unsupported file type: '{ext}'. Use .xlsx, .xls, or .csv")

        df = df.fillna("").reset_index(drop=True)
        df.replace("", pd.NA, inplace=True)
        df.dropna(how="all", inplace=True)
        df.dropna(axis=1, how="all", inplace=True)
        df.fillna("", inplace=True)

        stats = {
            "rows":   len(df),
            "cols":   len(df.columns),
            "size":   size_s,
            "sheets": sheets,
            "file":   os.path.basename(path),
            "ext":    ext,
        }
        logger.success(
            f"Import complete → {stats['rows']:,} rows × {stats['cols']} columns  ({size_s})"
        )
        return df, stats


class ColumnDetector:
    @staticmethod
    def auto_map_targets(df: pd.DataFrame, targets: List[str], logger: Logger) -> Dict[str, str]:
        """Attempt to auto-map a list of ERP target fields to source columns.

        Returns mapping: target -> source_column (string), or '' if not found.
        """
        cols = list(df.columns)
        cols_norm = {c: re.sub(r"[^a-z0-9]", "", c.lower()) for c in cols}
        mapping: Dict[str, str] = {}

        for target in targets:
            t_norm = re.sub(r"[^a-z0-9]", "", target.lower())

            # 1) exact normalized match
            for c, cn in cols_norm.items():
                if cn == t_norm:
                    mapping[target] = c
                    logger.success(f"Mapped '{target}' ← '{c}' (exact)")
                    break
            if target in mapping:
                continue

            # 2) check known aliases
            aliases = COLUMN_ALIASES.get(target, [])
            for alias in aliases:
                a_norm = re.sub(r"[^a-z0-9]", "", alias.lower())
                for c, cn in cols_norm.items():
                    if a_norm == cn or a_norm in cn or cn in a_norm:
                        mapping[target] = c
                        logger.info(f"Fuzzy-mapped '{target}' ← '{c}' (alias '{alias}')")
                        break
                if target in mapping:
                    break
            if target in mapping:
                continue

            # 3) substring match
            for c, cn in cols_norm.items():
                if t_norm in cn or cn in t_norm:
                    mapping[target] = c
                    logger.info(f"Fuzzy-mapped '{target}' ← '{c}' (substring)")
                    break
            if target in mapping:
                continue

            # 4) common heuristics for Section/Class
            if t_norm in ("section", "sec"):
                for c in cols:
                    if re.search(r"section|sec|section_name|class_section", c, re.IGNORECASE):
                        mapping[target] = c
                        logger.info(f"Auto-detected '{target}' ← '{c}'")
                        break

            if target not in mapping:
                mapping[target] = ""
                logger.warn(f"No match for required field: '{target}'")

        return mapping


class DateFormatDetector:
    @staticmethod
    def detect_formats(series: pd.Series, max_samples: int = 300) -> List[str]:
        """Return a list of suggested strptime formats (and 'Auto') ordered by frequency."""
        samples = list(series.dropna().astype(str).map(str.strip))[:max_samples]
        if not samples:
            return []

        patterns = [
            "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y", "%Y-%m-%d",
            "%d %b %Y", "%d %B %Y", "%m%d%Y", "%d%m%Y", "%Y%m%d", "%y/%m/%d",
            "%d-%b-%Y",
        ]
        counts: Dict[str, int] = {p: 0 for p in patterns}
        auto_count = 0

        for s in samples:
            s_clean = s
            parsed = False
            for p in patterns:
                try:
                    datetime.strptime(s_clean, p)
                    counts[p] += 1
                    parsed = True
                    break
                except Exception:
                    pass
            if not parsed:
                try:
                    pd.to_datetime(s_clean, dayfirst=True)
                    auto_count += 1
                except Exception:
                    try:
                        pd.to_datetime(s_clean, dayfirst=False)
                        auto_count += 1
                    except Exception:
                        pass

        choices: List[Tuple[int, str]] = []
        if auto_count:
            choices.append((auto_count, "Auto"))
        for p, c in counts.items():
            if c:
                choices.append((c, p))

        choices.sort(reverse=True)
        return [c for _, c in choices]

    @staticmethod
    def looks_like_date_column(column_name: str, series: pd.Series, max_samples: int = 40) -> bool:
        """Reject numeric-only columns that are unlikely to be DOB fields."""
        name = str(column_name).lower()
        if any(token in name for token in ("dob", "birth", "date")):
            return True

        samples = [str(v).strip() for v in series.dropna().astype(str).head(max_samples)]
        if not samples:
            return False

        numeric_only = 0
        dateish = 0
        for value in samples:
            if re.fullmatch(r"\d+", value):
                numeric_only += 1
                continue
            if re.search(r"[/\-.]", value):
                dateish += 1
                continue
            if re.search(r"\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec|january|february|march|april|june|july|august|september|october|november|december)\b", value, re.IGNORECASE):
                dateish += 1
                continue
            if re.fullmatch(r"\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4}", value):
                dateish += 1

        # If the column is mostly bare numbers, do not auto-suggest it as DOB.
        if numeric_only >= max(3, len(samples) * 0.7):
            return False

        return dateish >= max(2, len(samples) // 4)


class DataCleaner:
    @staticmethod
    def clean(
        df: pd.DataFrame,
        mapping: Dict[str, str],
        logger: Logger,
        progress_cb=None,
        keep_spaces: bool = True,
        dob_source_formats: Optional[Dict[str, str]] = None,
        dob_desired_formats: Optional[Dict[str, str]] = None,
    ) -> Tuple[pd.DataFrame, dict]:

        original_count = len(df)
        logger.info(f"Starting data cleaning on {original_count:,} rows …")

        rev_map = {v: k for k, v in mapping.items()}
        df = df.rename(columns=rev_map)
        keep = [c for c in ERP_MAPPING_FIELDS if c in df.columns]
        df = df[keep].copy()

        if progress_cb: progress_cb(0.1)

        for col in df.columns:
            df[col] = df[col].astype(str).str.strip()

        df.replace("", pd.NA, inplace=True)
        df.dropna(how="all", inplace=True)
        df.fillna("", inplace=True)
        after_empty = len(df)
        logger.info(f"Removed {original_count - after_empty:,} empty rows")

        if progress_cb: progress_cb(0.2)

        before_dup = len(df)
        df.drop_duplicates(inplace=True)
        dup_removed = before_dup - len(df)
        logger.info(f"Removed {dup_removed:,} duplicate rows")

        if progress_cb: progress_cb(0.35)

        for col in ("FullName", "Father Name", "Mother Name"):
            if col in df.columns:
                df[col] = df[col].apply(
                    lambda x: DataCleaner._clean_person_name(x)
                )

        if "Gender" in df.columns:
            GENDER_MAP = {
                "male": "Male", "m": "Male", "1": "Male",
                "female": "Female", "f": "Female", "2": "Female", "girl": "Female", "boy": "Male",
                "other": "Others", "others": "Others", "o": "Others", "3": "Others",
            }
            df["Gender"] = (
                df["Gender"].str.lower()
                .map(lambda x: GENDER_MAP.get(x, x.title() if x else ""))
            )

        if progress_cb: progress_cb(0.5)

        if "Guardian Contact Number" in df.columns:
            def _clean_guardian_phone(x):
                s = str(x) if x else ""
                s = re.sub(r"[^\d+\-]", "", s)[:15]
                digits = re.sub(r"\D", "", s)
                # treat short digit sequences as invalid → empty string
                return s if len(digits) >= 7 else ""

            df["Guardian Contact Number"] = df["Guardian Contact Number"].apply(_clean_guardian_phone)

        if "CurrentClass" in df.columns:
            df["CurrentClass"] = df["CurrentClass"].apply(DataCleaner._norm_class)

        # Ensure address contains only alphabets and numerics (remove punctuation/spaces)
        if "Permanent Address" in df.columns:
            def _clean_address(value):
                text = str(value).strip()
                if not text or text.lower() in ("nan", "none"):
                    return ""

                text = re.sub(
                    r"\b(?:temporary\s+address|temp\s+address|temporary|temp)\b.*$",
                    "",
                    text,
                    flags=re.IGNORECASE,
                )
                text = re.sub(
                    r"^\s*(?:permanent\s+address|present\s+address|address)\s*[:\-]?\s*",
                    "",
                    text,
                    flags=re.IGNORECASE,
                )

                if keep_spaces:
                    text = re.sub(r"[^A-Za-z0-9\s]", "", text)
                    text = re.sub(r"\s+", " ", text).strip()
                else:
                    text = re.sub(r"[^A-Za-z0-9]", "", text)
                    text = re.sub(r"\s+", "", text)

                return text.strip()

            if keep_spaces:
                df["Permanent Address"] = df["Permanent Address"].apply(_clean_address)
            else:
                df["Permanent Address"] = df["Permanent Address"].apply(_clean_address)

        if progress_cb: progress_cb(0.7)

        if "DOB" in df.columns or "Date_of_birth" in df.columns or "DOB" in mapping:
            # determine which column holds raw DOB values (source column)
            src_col = mapping.get("DOB") or mapping.get("Date_of_birth") or mapping.get("DOB")

            selected_source_fmt = None
            selected_desired_fmt = None
            if src_col:
                if dob_source_formats:
                    selected_source_fmt = dob_source_formats.get(src_col)
                if dob_desired_formats:
                    selected_desired_fmt = dob_desired_formats.get(src_col)

            def _parse_wrapper(v):
                return DataCleaner._parse_dob_with_format(
                    v,
                    source_fmt=selected_source_fmt,
                    desired_fmt=selected_desired_fmt,
                )

            # normalize whichever ERP-named column exists
            if "DOB" in df.columns:
                df["DOB"] = df["DOB"].apply(_parse_wrapper)
            elif "Date_of_birth" in df.columns:
                df["Date_of_birth"] = df["Date_of_birth"].apply(_parse_wrapper)

        # Ensure DOB column contains '0' where missing (user requirement)
        for _dob_col in ("DOB", "Date_of_birth"):
            if _dob_col in df.columns:
                df[_dob_col] = df[_dob_col].apply(lambda v: v if v else "0")

        invalid_mask  = pd.Series(False, index=df.index)
        if "FullName" in df.columns:
            invalid_mask |= df["FullName"].eq("")
        invalid_count = invalid_mask.sum()
        df = df[~invalid_mask].reset_index(drop=True)

        if progress_cb: progress_cb(1.0)

        stats = {
            "original":          original_count,
            "final":             len(df),
            "duplicates_removed": dup_removed,
            "invalid_removed":    invalid_count,
        }
        logger.success(
            f"Cleaning done → {len(df):,} valid rows  "
            f"(removed {dup_removed} duplicates, {invalid_count} invalid)"
        )
        return df, stats

    # Known word-only class names — kept as a plain class variable (no decorator)
    _WORD_CLASSES = {
        "NURSERY", "PLAY GROUP", "PLAYGROUP", "PLAY", "PRE-NURSERY",
        "PRE NURSERY", "PRENURSERY", "LKG", "UKG", "KG", "KINDERGARTEN",
        "PREP", "RECEPTION", "FOUNDATION", "PRE-PRIMARY", "PRE PRIMARY",
        "PREPRIMARY", "JUNIOR", "SENIOR", "MONTESSORI",
    }

    @staticmethod
    def _norm_class(val: str) -> str:
        if not val:
            return ""
        val = str(val).strip()
        u = val.upper().strip()

        # 1. Known word-only class names (Nursery, KG, LKG …) — preserve as Title Case
        if u in DataCleaner._WORD_CLASSES:
            return val.title()

        # 2. Roman numeral only (e.g. "VII", "XII") → "Class 7"
        ROMAN = {
            "I": "1", "II": "2", "III": "3", "IV": "4", "V": "5",
            "VI": "6", "VII": "7", "VIII": "8", "IX": "9", "X": "10",
            "XI": "11", "XII": "12",
        }
        if u in ROMAN:
            return f"Class {ROMAN[u]}"

        # 3. Pure digit (e.g. "7", "10") → "Class 7"
        if re.fullmatch(r"\d+", val.strip()):
            return f"Class {val.strip()}"

        # 4. "Class 7", "Grade 10", "Std 5" etc. — normalise to "Class N"
        if re.match(r"(?:class|grade|std|standard)\s*\d+", val, re.IGNORECASE):
            num = re.search(r"\d+", val).group()
            return f"Class {num}"

        # 5. Mixed alphanumeric like "KG1", "Prep2", "UKG2" — word part + number
        m = re.match(r"([A-Za-z][A-Za-z\s\-]*)(\d+)$", val.strip())
        if m:
            word_part = m.group(1).strip().title()
            num_part  = m.group(2)
            return f"{word_part} {num_part}"

        # 6. Purely alphabetic / hyphenated (unknown word class) — Title Case
        if re.fullmatch(r"[A-Za-z][A-Za-z\s\-]*", val):
            return val.title()

        # 7. Fallback — return original stripped value as-is
        return val

    @staticmethod
    def _clean_person_name(val: str) -> str:
        if not val:
            return ""
        cleaned = re.sub(r"[^A-Za-z\s.]", " ", str(val))
        cleaned = re.sub(r"\s+", " ", cleaned).strip()
        cleaned = re.sub(r"\s*\.\s*", ".", cleaned)
        return cleaned.title()

    @staticmethod
    def _email_name_token(val: str) -> str:
        cleaned = DataCleaner._clean_person_name(val)
        if not cleaned:
            return "student"
        tokens = [t for t in re.split(r"[\s.]+", cleaned) if t]
        if not tokens:
            return "student"
        base = re.sub(r"[^A-Za-z]", "", tokens[0]).lower()
        if len(base) < 3:
            for extra in tokens[1:]:
                base += re.sub(r"[^A-Za-z]", "", extra).lower()
                if len(base) >= 3:
                    break
        if len(base) < 3:
            joined = re.sub(r"[^A-Za-z]", "", cleaned).lower()
            if joined:
                base = joined
        return base or "student"

    @staticmethod
    def _class_email_token(val: str) -> str:
        cls = DataCleaner._norm_class(val)
        cls = cls.replace("Class ", "class")
        cls = re.sub(r"[^A-Za-z0-9]+", "", cls).lower()
        return cls or "class"

    @staticmethod
    def _parse_dob(val: str) -> str:
        # legacy signature kept for compatibility
        return DataCleaner._parse_dob_with_format(val, source_fmt=None, desired_fmt=None)

    @staticmethod
    def _parse_dob_with_format(
        val: str,
        source_fmt: Optional[str] = None,
        desired_fmt: Optional[str] = None,
    ) -> str:
        if not val or str(val).lower() in ("nan", "none", ""):
            return ""
        val = str(val).strip()
        dt = _parse_date_by_label(val, source_fmt)

        if dt is None:
            # fallback heuristics when no explicit source format works
            common = ["%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y", "%Y-%m-%d", "%d %b %Y", "%d %B %Y", "%d%m%Y", "%m%d%Y", "%Y%m%d", "%d-%b-%Y"]
            for fmt_try in common:
                try:
                    dt = datetime.strptime(val, fmt_try)
                    break
                except Exception:
                    pass

        if dt is None:
            # try pandas parser with dayfirst heuristics
            try:
                parsed = pd.to_datetime(val, dayfirst=True, errors="coerce")
                if pd.notnull(parsed):
                    dt = parsed.to_pydatetime()
            except Exception:
                pass
        if dt is None:
            try:
                parsed = pd.to_datetime(val, dayfirst=False, errors="coerce")
                if pd.notnull(parsed):
                    dt = parsed.to_pydatetime()
            except Exception:
                pass

        # Excel serial date (common when import from Excel yields numbers)
        if dt is None and re.fullmatch(r"\d{4,5}", val):
            try:
                days = int(val)
                dt = (datetime(1899, 12, 30) + pd.to_timedelta(days, unit="d")).to_pydatetime()
            except Exception:
                dt = None

        if dt is None:
            stripped = re.sub(r"[^0-9]", "", val)
            return stripped[:8] if len(stripped) >= 8 else (stripped if stripped else "")

        desired = desired_fmt if desired_fmt and desired_fmt != "Auto" else "YYYYMMDD"
        return _format_datetime_by_label(dt, desired)


class FormulaEngine:
    @staticmethod
    def build_erp(
        df: pd.DataFrame,
        app_state: "AppState",
        logger: Logger,
        progress_cb=None,
    ) -> pd.DataFrame:

        logger.info(f"Building ERP for {len(df):,} records …")
        domain = app_state.school_domain or "school.com"
        rows: List[List[Any]] = []

        def g(row, field: str) -> str:
            safe = field.replace(" ", "_").replace("'", "").replace("-", "_")
            return str(getattr(row, safe, "") or "").strip()

        def build_email(name_val: str, class_val: str, row_idx: int) -> str:
            name_token = DataCleaner._email_name_token(name_val)
            class_token = DataCleaner._class_email_token(class_val)
            suffix = f"{row_idx - 1:02d}"
            return f"{name_token}{class_token}{suffix}@{domain}"

        total = len(df)
        for idx, row in enumerate(df.itertuples(index=False), start=2):
            name     = g(row, "FullName")
            gender_r = g(row, "Gender")
            father   = g(row, "Father_Name")
            mother   = g(row, "Mother_Name")
            phone    = g(row, "Guardian_Contact_Number")
            dob      = g(row, "DOB")
            address  = g(row, "Permanent_Address")
            class_val = g(row, "CurrentClass") or g(row, "Class")

            gender_num = {"Male": 1, "Female": 2, "Others": 3}.get(gender_r, "")

            email = build_email(name, class_val, idx)

            guardian = father if father else mother
            relation = "Father" if father else ("Mother" if mother else "")
            g_email  = build_email(guardian or "guardian", class_val, idx)

            acad_raw = (app_state.academic_year or "=1").lstrip("=")
            try:
                acad_val: Any = int(acad_raw)
            except ValueError:
                acad_val = acad_raw

            # Determine Regd_no value (use existing column if configured)
            regd_val = ""
            if app_state.regd_no_mode == "Use Existing Column" and getattr(app_state, "regd_no_column", ""):
                regd_val = g(row, app_state.regd_no_column)

            rows.append([
                regd_val,
                name,
                email,
                gender_num,
                father,
                mother,
                phone,
                dob,
                address,
                acad_val,
                1,
                guardian,
                relation,
                g_email,
                address,
                phone,
                str(idx - 1),
                "",
                2,
            ])

            if progress_cb and idx % 500 == 0:
                progress_cb(idx / total)

        erp_df = pd.DataFrame(rows, columns=ERP_HEADERS)
        if progress_cb:
            progress_cb(1.0)
        logger.success(
            f"ERP structure ready → {len(erp_df):,} records × {len(ERP_HEADERS)} columns"
        )
        return erp_df


class ValidationEngine:
    @staticmethod
    def validate(df: pd.DataFrame) -> List[str]:
        issues: List[str] = []
        if "Name" in df.columns:
            missing_name = df["Name"].eq("").sum()
            if missing_name:
                issues.append(f"⚠️  {missing_name} rows missing student name")
        if "Date_of_birth" in df.columns:
            invalid_dob = df["Date_of_birth"].apply(
                lambda x: bool(x) and (len(str(x)) != 8 or not str(x).isdigit())
            ).sum()
            if invalid_dob:
                issues.append(f"⚠️  {invalid_dob} rows with non-standard DOB format")
        if "Phone_Number" in df.columns:
            short_phone = df["Phone_Number"].apply(
                lambda x: bool(x) and len(re.sub(r"\D", "", str(x))) < 7
            ).sum()
            if short_phone:
                issues.append(f"⚠️  {short_phone} rows with short/invalid phone numbers")
        if not issues:
            issues.append("✅  No critical validation issues found")
        return issues


class ExportEngine:
    @staticmethod
    def _class_filename_token(cls: str) -> str:
        cls = str(cls).strip()
        if not cls:
            return "Class"

        roman_map = {
            "I": "1", "II": "2", "III": "3", "IV": "4", "V": "5",
            "VI": "6", "VII": "7", "VIII": "8", "IX": "9", "X": "10",
            "XI": "11", "XII": "12",
        }

        if cls.upper() in roman_map:
            return f"Class_{roman_map[cls.upper()]}"

        if re.fullmatch(r"\d+", cls):
            return f"Class_{cls}"

        if re.fullmatch(r"[A-Za-z]+", cls):
            return f"Class_{cls.title()}"

        cls_safe = re.sub(r"[^\w]+", "_", cls).strip("_")
        return f"Class_{cls_safe or 'Class'}"

    @staticmethod
    def _safe_sheet_title(name: str) -> str:
        title = re.sub(r"[\\/*?:\[\]]", "_", str(name).strip())
        return title[:31] or "Sheet1"

    @staticmethod
    def _write_xlsx(df: pd.DataFrame, path: str, sheet_name: str, app_state: "AppState"):
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("Unable to initialize Excel worksheet")
        ws.title = ExportEngine._safe_sheet_title(sheet_name)

        H_FILL  = PatternFill("solid", fgColor="4C4CA0")
        H_FONT  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
        BORDER  = Border(bottom=Side(style="thin", color="7070CC"))

        for c_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=c_idx, value=col)
            cell.fill  = H_FILL
            cell.font  = H_FONT
            cell.alignment = H_ALIGN
            cell.border    = BORDER

        EVEN_FILL = PatternFill("solid", fgColor="EDEDF9")
        ODD_FILL  = PatternFill("solid", fgColor="FFFFFF")
        D_ALIGN   = Alignment(horizontal="left", vertical="center")
        D_FONT    = Font(name="Calibri", size=10)

        for r_idx, row_data in enumerate(df.itertuples(index=False), start=2):
            fill = EVEN_FILL if r_idx % 2 == 0 else ODD_FILL
            for c_idx, _ in enumerate(df.columns, 1):
                col_name = df.columns[c_idx - 1]
                val      = row_data[c_idx - 1]
                cell     = ws.cell(row=r_idx, column=c_idx)

                if col_name == "Regd_no":
                    mode = app_state.regd_no_mode
                    if mode == "=ROW()-1":
                        cell.value = f"=ROW()-1"
                    elif mode == '="000"&ROW()-1':
                        cell.value = f'="000"&ROW()-1'
                    elif mode == "Custom Formula":
                        cell.value = app_state.regd_no_custom or "=ROW()-1"
                    else:
                        cell.value = val or f"=ROW()-1"

                elif col_name == "roll_number":
                    cell.value = f"=ROW()-1"

                elif col_name == "Email":
                    cell.value = val

                elif col_name == "Guardian_email":
                    cell.value = val
                else:
                    cell.value = (val if val != "" else None)

                cell.fill      = fill
                cell.font      = D_FONT
                cell.alignment = D_ALIGN

        for col in ws.columns:
            letter  = col[0].column_letter
            max_len = max(
                (len(str(c.value)) for c in col if c.value is not None), default=10
            )
            ws.column_dimensions[letter].width = min(max(max_len + 3, 12), 38)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 28

        wb.save(path)

    @staticmethod
    def export_all(
        groups: Dict[str, pd.DataFrame],
        app_state: "AppState",
        logger: Logger,
        progress_cb=None,
    ) -> List[Tuple[str, str]]:
        """Returns list of (filepath, original_class_key) tuples."""

        folder = app_state.output_folder
        os.makedirs(folder, exist_ok=True)
        exported: List[Tuple[str, str]] = []
        total = len(groups)

        school_safe = re.sub(r"[^\w]", "_", app_state.school_name or "School")

        # Track used filenames to avoid collisions between classes that
        # normalise to the same safe name (e.g. "KG 1" and "KG-1").
        used_names: Dict[str, int] = {}

        for i, (cls, df) in enumerate(groups.items()):
            cls_safe = ExportEngine._class_filename_token(cls)

            ext   = app_state.output_format.lower()
            base  = f"{school_safe}_{cls_safe}"

            # Deduplicate if two classes produce the same safe name
            if base in used_names:
                used_names[base] += 1
                base = f"{base}_{used_names[base]}"
            else:
                used_names[base] = 0

            fname = f"{base}.{ext}"
            fpath = os.path.join(folder, fname)

            export_df = df.copy().reset_index(drop=True)
            export_df["roll_number"] = range(1, len(export_df) + 1)

            try:
                if app_state.output_format == "XLSX":
                    ExportEngine._write_xlsx(export_df, fpath, cls, app_state)
                else:
                    export_df.to_csv(fpath, index=False, encoding="utf-8-sig")

                exported.append((fpath, cls))
                logger.success(f"Exported: {fname}  ({len(export_df):,} rows)")
            except Exception as e:
                logger.error(f"Failed to export {fname}: {e}")

            if progress_cb:
                progress_cb((i + 1) / total)

        return exported


class PreviewTable(ctk.CTkFrame):
    def __init__(self, parent, **kwargs):
        super().__init__(parent, fg_color="transparent", **kwargs)
        self._df: Optional[pd.DataFrame] = None
        self._max_rows: int = 15
        self._sort_state: Dict[str, bool] = {}
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", self._on_search)
        self._build_ui()

    def _build_ui(self):
        top = ctk.CTkFrame(self, fg_color="transparent", height=36)
        top.pack(fill="x", padx=4, pady=(4, 2))
        top.pack_propagate(False)

        ctk.CTkEntry(
            top, textvariable=self._search_var,
            placeholder_text="🔍 Search rows...",
            width=240, height=30, corner_radius=8,
        ).pack(side="left", padx=(0, 10))

        self._lbl_stats = ctk.CTkLabel(
            top, text="", text_color=("gray40", "gray60"),
            font=ctk.CTkFont(size=11),
        )
        self._lbl_stats.pack(side="right")

        tree_wrap = ctk.CTkFrame(self, corner_radius=8)
        tree_wrap.pack(fill="both", expand=True, padx=4, pady=(0, 4))
        tree_wrap.grid_rowconfigure(0, weight=1)
        tree_wrap.grid_columnconfigure(0, weight=1)

        self._apply_treeview_style()
        self._tree = ttk.Treeview(
            tree_wrap, style="ERP.Treeview", show="headings", selectmode="browse"
        )

        vsb = ttk.Scrollbar(tree_wrap, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(tree_wrap, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        self._tree.tag_configure("odd",  background="#21253A", foreground="#E0E0E0")
        self._tree.tag_configure("even", background="#1A1D27", foreground="#E0E0E0")

    @staticmethod
    def _apply_treeview_style():
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "ERP.Treeview",
            background="#1A1D27", foreground="#E0E0E0",
            rowheight=26, fieldbackground="#1A1D27",
            borderwidth=0, font=("Segoe UI", 10),
        )
        style.configure(
            "ERP.Treeview.Heading",
            background="#2D3250", foreground="#FFFFFF",
            font=("Segoe UI", 10, "bold"), borderwidth=0,
            relief="flat",
        )
        style.map(
            "ERP.Treeview",
            background=[("selected", "#6C63FF")],
            foreground=[("selected", "#FFFFFF")],
        )
        style.map(
            "ERP.Treeview.Heading",
            background=[("active", "#3D4270")],
        )

    def load_dataframe(self, df: pd.DataFrame, max_rows: int = 15):
        self._df       = df.reset_index(drop=True)
        self._max_rows = max_rows
        self._render(self._df.head(max_rows))

    def _render(self, df: pd.DataFrame):
        self._tree.delete(*self._tree.get_children())
        if df is None or df.empty:
            self._tree["columns"] = []
            self._lbl_stats.configure(text="No data to display")
            return

        cols = list(df.columns)
        self._tree["columns"] = cols
        for col in cols:
            self._tree.heading(col, text=col, command=lambda c=col: self._sort(c))
            width = max(len(col) * 9, 72)
            self._tree.column(col, width=min(width, 190), minwidth=60, anchor="w")

        for i, (_, row) in enumerate(df.iterrows()):
            tag  = "odd" if i % 2 else "even"
            vals = [str(v)[:80] for v in row]
            self._tree.insert("", "end", values=vals, tags=(tag,))

        total  = len(self._df) if self._df is not None else len(df)
        shown  = len(df)
        n_cols = len(cols)
        self._lbl_stats.configure(
            text=f"Showing {shown:,} of {total:,} rows  ·  {n_cols} columns"
        )

    def _on_search(self, *_):
        if self._df is None:
            return
        q = self._search_var.get().strip().lower()
        if not q:
            self._render(self._df.head(self._max_rows))
            return
        mask    = self._df.astype(str).apply(lambda r: r.str.lower().str.contains(q, na=False).any(), axis=1)
        results = self._df[mask].head(100)
        self._render(results)

    def _sort(self, col: str):
        if self._df is None or col not in self._df.columns:
            return
        asc = not self._sort_state.get(col, True)
        self._sort_state[col] = asc
        sorted_df = self._df.sort_values(col, ascending=asc, na_position="last")
        self._render(sorted_df.head(self._max_rows))


class LogPanel(ctk.CTkFrame):
    LEVEL_PREFIXES = {
        "INFO":  "ℹ️  ",
        "OK":    "✅  ",
        "WARN":  "⚠️  ",
        "ERROR": "❌  ",
    }

    def __init__(self, parent, **kwargs):
        super().__init__(parent, **kwargs)
        self._build_ui()

    def _build_ui(self):
        hdr = ctk.CTkFrame(self, fg_color="transparent", height=28)
        hdr.pack(fill="x", padx=8, pady=(6, 0))
        hdr.pack_propagate(False)

        ctk.CTkLabel(
            hdr, text="📋  Live Logs",
            font=ctk.CTkFont(size=11, weight="bold"),
        ).pack(side="left")

        ctk.CTkButton(
            hdr, text="Clear", width=48, height=22,
            corner_radius=4, fg_color="transparent",
            border_width=1, border_color="gray50",
            command=self._clear,
        ).pack(side="right")

        self._box = ctk.CTkTextbox(
            self,
            font=ctk.CTkFont(family="Consolas", size=10),
            corner_radius=6, border_width=0,
            state="disabled",
        )
        self._box.pack(fill="both", expand=True, padx=8, pady=(4, 8))

    def append(self, entry: dict):
        prefix = self.LEVEL_PREFIXES.get(entry.get("level", "INFO"), "   ")
        line   = f"[{entry['ts']}] {prefix}{entry['msg']}\n"
        self._box.configure(state="normal")
        self._box.insert("end", line)
        self._box.configure(state="disabled")
        self._box.see("end")

    def _clear(self):
        self._box.configure(state="normal")
        self._box.delete("1.0", "end")
        self._box.configure(state="disabled")


class StatCard(ctk.CTkFrame):
    def __init__(self, parent, title: str, value: str,
                 icon: str, color: str = "#6C63FF", **kwargs):
        super().__init__(parent, corner_radius=12, **kwargs)

        ctk.CTkLabel(self, text=icon, font=ctk.CTkFont(size=26)).pack(pady=(14, 2))
        ctk.CTkLabel(
            self, text=value,
            font=ctk.CTkFont(size=22, weight="bold"),
            text_color=color,
        ).pack(pady=(0, 2))
        ctk.CTkLabel(
            self, text=title,
            font=ctk.CTkFont(size=11),
            text_color=("gray45", "gray55"),
        ).pack(pady=(0, 14))


# ══════════════════════════════════════════════════════════════════════════════
# MAIN APPLICATION  — self.state renamed to self.app_state to avoid
#                     conflict with tkinter's built-in CTk.state() method
# ══════════════════════════════════════════════════════════════════════════════

class SchoolERPApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # ── FIX: use app_state, not state (clashes with CTk.state()) ──
        self.app_state = AppState()

        self._step_frames: Dict[int, ctk.CTkFrame] = {}
        self._step_btns:   Dict[int, ctk.CTkButton] = {}
        self._current_step: int = 1

        ctk.set_default_color_theme("blue")

        self.title(f"🏫  {APP_NAME}  v{APP_VERSION}")
        self.geometry(WIN_SIZE)
        self.minsize(1100, 700)
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        self._build_layout()
        self._build_sidebar()
        self._build_all_steps()
        self._show_step(1)

        # maximize after UI is built to avoid later layout calls restoring size
        def _do_maximize():
            try:
                self.state("zoomed")
            except Exception:
                try:
                    self.attributes("-zoomed", True)
                except Exception:
                    pass

        try:
            self.after(150, _do_maximize)
        except Exception:
            _do_maximize()

    def _build_layout(self):
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._sidebar = ctk.CTkFrame(
            self, width=230, corner_radius=0,
            fg_color=("#ECEEFF", "#13162A"),
        )
        self._sidebar.grid(row=0, column=0, sticky="nsew")
        self._sidebar.grid_propagate(False)

        self._main = ctk.CTkFrame(
            self, corner_radius=0,
            fg_color=("gray94", "#0F1117"),
        )
        self._main.grid(row=0, column=1, sticky="nsew")
        self._main.grid_rowconfigure(0, weight=1)
        self._main.grid_columnconfigure(0, weight=1)

    def _build_sidebar(self):
        sb = self._sidebar

        logo = ctk.CTkFrame(sb, fg_color="transparent", height=76)
        logo.pack(fill="x", pady=(14, 4))
        logo.pack_propagate(False)
        ctk.CTkLabel(logo, text="🏫", font=ctk.CTkFont(size=32)).pack(pady=(6, 0))
        ctk.CTkLabel(
            logo, text="School ERP",
            font=ctk.CTkFont(size=13, weight="bold"),
        ).pack()

        _divider(sb)

        for step_id, label in STEPS:
            btn = ctk.CTkButton(
                sb, text=label, anchor="w",
                height=42, corner_radius=10,
                fg_color="transparent",
                hover_color=("#D5D8FF", "#252843"),
                text_color=("gray30", "gray65"),
                font=ctk.CTkFont(size=12),
                command=lambda s=step_id: self._try_navigate(s),
            )
            btn.pack(fill="x", padx=10, pady=2)
            self._step_btns[step_id] = btn

        _divider(sb)

        # theme and undo removed per user request

        ctk.CTkButton(
            sb, text="📋  Recent Files", height=34, corner_radius=8,
            fg_color="transparent",
            hover_color=("#D5D8FF", "#252843"),
            command=self._show_recent,
        ).pack(fill="x", padx=10, pady=3)

        ctk.CTkLabel(
            sb, text=f"v{APP_VERSION}",
            font=ctk.CTkFont(size=10),
            text_color="gray50",
        ).pack(side="bottom", pady=8)

    def _toggle_theme(self):
        # theme toggle removed
        return

    def _undo(self):
        # undo feature removed
        messagebox.showinfo("Undo", "Undo feature is disabled.")

    def _show_recent(self):
        recent = AppState.get_recent()
        if not recent:
            messagebox.showinfo("Recent Files", "No recent files found.")
            return
        win = ctk.CTkToplevel(self)
        win.title("Recent Files")
        win.geometry("500x300")
        win.grab_set()
        ctk.CTkLabel(win, text="Recent Files", font=ctk.CTkFont(size=14, weight="bold")).pack(pady=12)
        for path in recent[:10]:
            def _open(p=path):
                win.destroy()
                self._load_file(p)
            ctk.CTkButton(
                win, text=os.path.basename(path),
                anchor="w", height=32, corner_radius=6,
                fg_color="transparent",
                hover_color=("#D5D8FF", "#252843"),
                command=_open,
            ).pack(fill="x", padx=16, pady=2)

    def _show_step(self, step_id: int):
        for f in self._step_frames.values():
            f.grid_remove()
        if step_id in self._step_frames:
            self._step_frames[step_id].grid(row=0, column=0, sticky="nsew")

        for sid, btn in self._step_btns.items():
            if sid == step_id:
                btn.configure(fg_color=("#6C63FF", "#6C63FF"), text_color="white")
            elif sid < step_id:
                btn.configure(fg_color="transparent", text_color=("#2E7D32", "#66BB6A"))
            else:
                btn.configure(fg_color="transparent", text_color=("gray45", "gray60"))

        self._current_step = step_id
        # Refresh dynamic UI elements for certain steps
        if step_id == 4:
            try:
                self._s4_update_regd_columns()
            except Exception:
                pass

    def _s4_update_regd_columns(self):
        cols = []
        # Prefer cleaned_df, fall back to raw_df
        try:
            if self.app_state.cleaned_df is not None:
                cols = list(self.app_state.cleaned_df.columns)
            elif self.app_state.raw_df is not None:
                cols = list(self.app_state.raw_df.columns)
        except Exception:
            cols = []
        # Populate combobox
        if hasattr(self, "_s4_regd_col_cb") and self._s4_regd_col_cb is not None:
            try:
                self._s4_regd_col_cb.configure(values=cols)
                if self.app_state.regd_no_column:
                    try:
                        self._s4_regd_col_var.set(self.app_state.regd_no_column)
                    except Exception:
                        pass
            except Exception:
                pass

    def _try_navigate(self, step_id: int):
        if step_id <= self._current_step:
            self._show_step(step_id)

    def _advance(self):
        nxt = self._current_step + 1
        if nxt <= 7:
            self._show_step(nxt)

    def _logger(self, panel: LogPanel) -> Logger:
        def cb(entry):
            self.after(0, lambda e=entry: panel.append(e))
        return Logger(cb)

    def _make_header(self, parent, title: str, subtitle: str = "") -> None:
        hdr = ctk.CTkFrame(
            parent, height=68, corner_radius=0,
            fg_color=("white", "#13162A"),
            border_width=0,
        )
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.grid_propagate(False)
        hdr.grid_columnconfigure(0, weight=1)
        hdr.grid_rowconfigure(0, weight=1)

        inner = ctk.CTkFrame(hdr, fg_color="transparent")
        inner.grid(row=0, column=0, sticky="w", padx=26)
        ctk.CTkLabel(inner, text=title, font=ctk.CTkFont(size=19, weight="bold")).pack(anchor="w")
        if subtitle:
            ctk.CTkLabel(
                inner, text=subtitle,
                text_color=("gray55", "gray60"),
                font=ctk.CTkFont(size=12),
            ).pack(anchor="w")

    def _make_footer(self, parent, cmd, label: str = "Move to Next Step →") -> None:
        ftr = ctk.CTkFrame(
            parent, height=62, corner_radius=0,
            fg_color=("white", "#13162A"),
        )
        ftr.grid(row=2, column=0, sticky="ew")
        ftr.grid_propagate(False)
        ftr.grid_columnconfigure(0, weight=1)
        ftr.grid_rowconfigure(0, weight=1)

        ctk.CTkButton(
            ftr, text=label,
            height=40, width=220, corner_radius=10,
            fg_color="#6C63FF", hover_color="#5550CC",
            font=ctk.CTkFont(size=13, weight="bold"),
            command=cmd,
        ).grid(row=0, column=0, sticky="e", padx=24)

    def _render_stat_cards(self, parent_frame: ctk.CTkFrame,
                           cards: List[Tuple[str, str, str, str]]):
        for w in parent_frame.winfo_children():
            w.destroy()
        for i, (title, value, icon, color) in enumerate(cards):
            card = StatCard(parent_frame, title, value, icon, color)
            card.grid(row=0, column=i, padx=6, pady=4, sticky="ew")
            parent_frame.grid_columnconfigure(i, weight=1)

    def _build_all_steps(self):
        self._build_step1()
        self._build_step2()
        self._build_step3()
        self._build_step4()
        self._build_step5()
        self._build_step6()
        self._build_step7()

    # ── STEP 1 ────────────────────────────────────────────────────────────

    def _build_step1(self):
        frame = self._new_step_frame()
        self._make_header(frame, "Step 1 — Import Student File",
                          "Supports XLSX • XLS • CSV  |  50,000+ rows")

        scroll = ctk.CTkScrollableFrame(frame, fg_color="transparent")
        scroll.grid(row=1, column=0, sticky="nsew", padx=22, pady=8)
        scroll.grid_columnconfigure(0, weight=1)

        drop = ctk.CTkFrame(scroll, corner_radius=16, height=162,
                            border_width=2, border_color="#6C63FF")
        drop.grid(row=0, column=0, sticky="ew", pady=(0, 14))
        drop.grid_propagate(False)
        drop.grid_columnconfigure(0, weight=1)
        drop.grid_rowconfigure(0, weight=1)

        inner = ctk.CTkFrame(drop, fg_color="transparent")
        inner.grid(row=0, column=0)
        ctk.CTkLabel(inner, text="📂", font=ctk.CTkFont(size=44)).pack()
        ctk.CTkLabel(inner, text="Click to Browse or Drag & Drop",
                     font=ctk.CTkFont(size=15, weight="bold")).pack(pady=(4, 2))
        ctk.CTkLabel(inner, text="XLSX  ·  XLS  ·  CSV   |   Up to 50,000+ students",
                     text_color="gray55").pack()
        ctk.CTkButton(inner, text="  Browse Files  ", height=38, width=150,
                      corner_radius=8, command=lambda: self._s1_browse()).pack(pady=(10, 0))

        drop.bind("<Button-1>", lambda _: self._s1_browse())
        inner.bind("<Button-1>", lambda _: self._s1_browse())

        self._s1_cards = ctk.CTkFrame(scroll, fg_color="transparent")
        self._s1_cards.grid(row=1, column=0, sticky="ew", pady=(0, 10))

        self._s1_prog = ctk.CTkProgressBar(scroll, height=5, corner_radius=3)
        self._s1_prog.grid(row=2, column=0, sticky="ew", pady=(0, 6))
        self._s1_prog.set(0)

        self._s1_prog_lbl = ctk.CTkLabel(scroll, text="", text_color="gray55",
                                          font=ctk.CTkFont(size=11))
        self._s1_prog_lbl.grid(row=3, column=0, sticky="w", pady=(0, 8))

        ctk.CTkLabel(scroll, text="Data Preview",
                     font=ctk.CTkFont(size=13, weight="bold")).grid(
            row=4, column=0, sticky="w", pady=(0, 4))

        self._s1_preview = PreviewTable(scroll)
        self._s1_preview.grid(row=5, column=0, sticky="nsew")
        scroll.grid_rowconfigure(5, weight=1)

        self._s1_log = LogPanel(scroll, height=110)
        self._s1_log.grid(row=6, column=0, sticky="ew", pady=(8, 0))

        self._make_footer(frame, self._s1_next)
        self._step_frames[1] = frame

    def _s1_browse(self):
        path = filedialog.askopenfilename(
            title="Select Student Data File",
            filetypes=[
                ("Supported formats", "*.xlsx *.xls *.csv"),
                ("Excel",  "*.xlsx *.xls"),
                ("CSV",    "*.csv"),
                ("All",    "*.*"),
            ],
        )
        if path:
            self._load_file(path)

    def _load_file(self, path: str):
        logger = self._logger(self._s1_log)
        self._s1_prog.set(0.1)
        self._s1_prog_lbl.configure(text="Loading file …")
        self._show_step(1)

        def work():
            try:
                df, stats = ImportEngine.load(path, logger)
                self.app_state.raw_df      = df
                self.app_state.source_file = path
                self.app_state.import_stats = stats
                AppState.add_recent(path)
                self.after(0, lambda: self._s1_done(stats, df))
            except Exception as exc:
                logger.error(str(exc))
                self.after(0, lambda: (
                    self._s1_prog_lbl.configure(text=f"❌ Error: {exc}", text_color="#F44336"),
                    messagebox.showerror("Import Error", str(exc)),
                ))

        threading.Thread(target=work, daemon=True).start()

    def _s1_done(self, stats: dict, df: pd.DataFrame):
        self._s1_prog.set(1.0)
        self._s1_prog_lbl.configure(text="✅  Import complete!", text_color="#4CAF50")
        self._render_stat_cards(self._s1_cards, [
            ("Total Rows",    f"{stats['rows']:,}",      "📊", "#6C63FF"),
            ("Columns",       str(stats["cols"]),        "📋", "#4CAF50"),
            ("File Size",     stats["size"],             "💾", "#FF9800"),
            ("Sheets Found",  str(len(stats["sheets"])), "📄", "#2196F3"),
        ])
        self._s1_preview.load_dataframe(df)

    def _s1_next(self):
        if self.app_state.raw_df is None:
            messagebox.showwarning("No File", "Please import a student data file first.")
            return
        self._advance()
        self._run_s2_detect()

    # ── STEP 2 ────────────────────────────────────────────────────────────

    def _build_step2(self):
        frame = self._new_step_frame()
        self._make_header(frame, "Step 2 — Auto Column Detection",
                          "Maps source columns to ERP fields intelligently")

        scroll = ctk.CTkScrollableFrame(frame, fg_color="transparent")
        scroll.grid(row=1, column=0, sticky="nsew", padx=22, pady=8)
        scroll.grid_columnconfigure(0, weight=1)
        scroll.grid_columnconfigure(1, weight=1)

        map_card = ctk.CTkFrame(scroll, corner_radius=12)
        map_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=(0, 10))
        map_card.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(map_card, text="🔍  Detected Column Mappings",
                     font=ctk.CTkFont(size=13, weight="bold")).grid(
            row=0, column=0, columnspan=3, sticky="w", padx=14, pady=(12, 6))

        self._s2_map_inner = ctk.CTkFrame(map_card, fg_color="transparent")
        self._s2_map_inner.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 12))
        self._s2_map_inner.grid_columnconfigure(1, weight=1)

        warn_card = ctk.CTkFrame(scroll, corner_radius=12)
        warn_card.grid(row=0, column=1, sticky="nsew", padx=(8, 0), pady=(0, 10))
        ctk.CTkLabel(warn_card, text="⚠️  Warnings & Issues",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(
            anchor="w", padx=14, pady=(12, 6))

        self._s2_warn_box = ctk.CTkTextbox(warn_card, height=180, corner_radius=6,
                                            state="disabled")
        self._s2_warn_box.pack(fill="x", padx=14, pady=(0, 12))

        ctk.CTkLabel(scroll, text="Mapped Data Preview",
                     font=ctk.CTkFont(size=13, weight="bold")).grid(
            row=1, column=0, columnspan=2, sticky="w", pady=(0, 4))

        self._s2_preview = PreviewTable(scroll)
        self._s2_preview.grid(row=2, column=0, columnspan=2, sticky="nsew")
        scroll.grid_rowconfigure(2, weight=1)

        self._s2_log = LogPanel(scroll, height=100)
        self._s2_log.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(8, 0))

        self._make_footer(frame, self._s2_next)
        self._step_frames[2] = frame
        self._s2_dropdowns: Dict[str, ctk.CTkComboBox] = {}

    def _run_s2_detect(self):
        logger = self._logger(self._s2_log)

        def work():
            mapping = ColumnDetector.auto_map_targets(self.app_state.raw_df, ERP_MAPPING_FIELDS, logger)
            # prevent duplicate source-column assignments (keep first occurrence)
            used: set = set()
            final: Dict[str, str] = {}
            for t, src in mapping.items():
                if src and src in used:
                    logger.warn(f"Duplicate mapping ignored for '{t}' → '{src}'")
                    final[t] = ""
                else:
                    final[t] = src
                    if src:
                        used.add(src)

            self.app_state.column_mapping = final
            self.after(0, lambda: self._s2_done(final, logger))

        threading.Thread(target=work, daemon=True).start()

    def _s2_done(self, mapping: Dict[str, str], logger: Logger):
        df = self.app_state.raw_df
        for w in self._s2_map_inner.winfo_children():
            w.destroy()
        self._s2_dropdowns.clear()

        all_cols = ["(not mapped)"] + list(df.columns)

        for r, target in enumerate(ERP_MAPPING_FIELDS):
            ctk.CTkLabel(
                self._s2_map_inner, text=target,
                font=ctk.CTkFont(size=11, weight="bold"),
                width=188, anchor="w",
            ).grid(row=r, column=0, sticky="w", pady=3)
            current = mapping.get(target, "(not mapped)")
            var = tk.StringVar(value=current)
            dd = ctk.CTkComboBox(
                self._s2_map_inner, values=all_cols,
                variable=var, width=220, height=28,
            )
            dd.grid(row=r, column=1, sticky="w", padx=(8, 0), pady=3)
            self._s2_dropdowns[target] = dd
            icon = "✅" if current else "❌"
            ctk.CTkLabel(self._s2_map_inner, text=icon, width=26).grid(
                row=r, column=2, padx=(4, 0))
        missing = [t for t, src in mapping.items() if not src]
        src_used = set(v for v in mapping.values() if v)
        unmapped = [c for c in df.columns if c not in src_used]
        lines: List[str] = []
        if missing:
            lines.append(f"⚠️  Missing ERP fields ({len(missing)}):")
            lines += [f"   • {m}" for m in missing]
        if unmapped:
            lines.append(f"\nℹ️  Unmapped source columns ({len(unmapped)}):")
            lines += [f"   • {u}" for u in unmapped[:12]]
        if not lines:
            lines = ["✅  All required fields detected successfully!"]

        self._s2_warn_box.configure(state="normal")
        self._s2_warn_box.delete("1.0", "end")
        self._s2_warn_box.insert("1.0", "\n".join(lines))
        self._s2_warn_box.configure(state="disabled")

        renamed = df.rename(columns={v: k for k, v in mapping.items() if v})
        keep = [c for c in ERP_MAPPING_FIELDS if c in renamed.columns]
        self._s2_preview.load_dataframe(renamed[keep] if keep else renamed)

        # Format customizer UI
        try:
            # remove existing dob frame
            if hasattr(self, "_s2_dob_frame") and self._s2_dob_frame is not None:
                self._s2_dob_frame.destroy()
        except Exception:
            pass

        self._s2_dob_frame = ctk.CTkFrame(self._s2_map_inner, fg_color="transparent")
        self._s2_dob_frame.grid(row=len(ERP_MAPPING_FIELDS) + 1, column=0, columnspan=3, sticky="ew", pady=(8,0))
        ctk.CTkLabel(self._s2_dob_frame, text="🛠️  Format Customizer",
                     font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, column=0, sticky="w", padx=4, pady=(4,4))
        ctk.CTkLabel(
            self._s2_dob_frame,
            text="Column",
            width=160,
            anchor="w",
            font=ctk.CTkFont(size=11, weight="bold"),
        ).grid(row=1, column=0, padx=6, pady=(0, 4))
        ctk.CTkLabel(
            self._s2_dob_frame,
            text="Select Current Format",
            width=260,
            anchor="w",
            font=ctk.CTkFont(size=11, weight="bold"),
        ).grid(row=1, column=1, padx=6, pady=(0, 4))
        ctk.CTkLabel(
            self._s2_dob_frame,
            text="Choose Desired Format",
            width=260,
            anchor="w",
            font=ctk.CTkFont(size=11, weight="bold"),
        ).grid(row=1, column=2, padx=(8, 0), pady=(0, 4))

        source_col = mapping.get("DOB") or mapping.get("Date_of_birth")
        if source_col and source_col in df.columns:
            row = 2
            ctk.CTkLabel(self._s2_dob_frame, text=source_col, width=160, anchor="w").grid(row=row, column=0, padx=6, pady=2)
            current_var = tk.StringVar(value=self.app_state.dob_column_format.get(source_col, "Auto"))
            desired_var = tk.StringVar(value=self.app_state.dob_desired_format.get(source_col, "Auto"))

            current_cb = ctk.CTkComboBox(
                self._s2_dob_frame,
                values=["Auto"] + DATE_FORMAT_OPTIONS,
                variable=current_var,
                width=260,
            )
            current_cb.grid(row=row, column=1, padx=6, pady=2)

            desired_cb = ctk.CTkComboBox(
                self._s2_dob_frame,
                values=["Auto"] + DATE_FORMAT_OPTIONS,
                variable=desired_var,
                width=260,
            )
            desired_cb.grid(row=row, column=2, padx=(8, 0), pady=2)

            def _sync_formats(event=None, col=source_col, current_var=current_var, desired_var=desired_var):
                self.app_state.dob_column_format[col] = current_var.get().strip() or "Auto"
                self.app_state.dob_desired_format[col] = desired_var.get().strip() or "Auto"

            current_cb.configure(command=_sync_formats)
            desired_cb.configure(command=_sync_formats)
            _sync_formats()
        else:
            ctk.CTkLabel(
                self._s2_dob_frame,
                text="DOB column not mapped; date format options are hidden.",
                text_color="gray50",
                anchor="w",
            ).grid(row=2, column=0, columnspan=3, padx=6, pady=2, sticky="w")

        address_row = 3
        ctk.CTkLabel(
            self._s2_dob_frame,
            text="Address Format",
            width=160,
            anchor="w",
            font=ctk.CTkFont(size=11, weight="bold"),
        ).grid(row=address_row, column=0, padx=6, pady=(12, 4), sticky="w")

        self._s2_addr_keep_var = tk.StringVar(value="Keep spaces" if self.app_state.address_keep_spaces else "Remove spaces")
        keep_rb = ctk.CTkRadioButton(
            self._s2_dob_frame,
            text="Keep spaces",
            variable=self._s2_addr_keep_var,
            value="Keep spaces",
            command=lambda: self._set_address_keep_spaces(True),
        )
        keep_rb.grid(row=address_row, column=1, padx=6, pady=(12, 4), sticky="w")

        remove_rb = ctk.CTkRadioButton(
            self._s2_dob_frame,
            text="Remove spaces",
            variable=self._s2_addr_keep_var,
            value="Remove spaces",
            command=lambda: self._set_address_keep_spaces(False),
        )
        remove_rb.grid(row=address_row, column=2, padx=(8, 0), pady=(12, 4), sticky="w")

        self._set_address_keep_spaces(self.app_state.address_keep_spaces)

    def _s2_next(self):
        mapping: Dict[str, str] = {}
        for target, dd in self._s2_dropdowns.items():
            val = dd.get()
            if val and val != "(not mapped)":
                mapping[target] = val
        if "FullName" not in mapping:
            messagebox.showwarning(
                "Required Field",
                "The 'FullName' field must be mapped before proceeding.",
            )
            return
        self.app_state.column_mapping = mapping
        self._advance()
        self._run_s3_clean()

    # ── STEP 3 ────────────────────────────────────────────────────────────

    def _build_step3(self):
        frame = self._new_step_frame()
        self._make_header(frame, "Step 3 — Data Cleaning",
                          "Remove duplicates, normalise values, and validate records")

        scroll = ctk.CTkScrollableFrame(frame, fg_color="transparent")
        scroll.grid(row=1, column=0, sticky="nsew", padx=22, pady=8)
        scroll.grid_columnconfigure(0, weight=1)

        self._s3_cards = ctk.CTkFrame(scroll, fg_color="transparent")
        self._s3_cards.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        self._s3_prog = ctk.CTkProgressBar(scroll, height=6, corner_radius=3)
        self._s3_prog.grid(row=1, column=0, sticky="ew", pady=(0, 4))
        self._s3_prog.set(0)

        self._s3_lbl = ctk.CTkLabel(scroll, text="Waiting …", text_color="gray55",
                                     font=ctk.CTkFont(size=11))
        self._s3_lbl.grid(row=2, column=0, sticky="w", pady=(0, 10))

        ctk.CTkLabel(scroll, text="Cleaned Data Preview",
                     font=ctk.CTkFont(size=13, weight="bold")).grid(
            row=3, column=0, sticky="w", pady=(0, 4))

        self._s3_preview = PreviewTable(scroll)
        self._s3_preview.grid(row=4, column=0, sticky="nsew")
        scroll.grid_rowconfigure(4, weight=1)

        self._s3_log = LogPanel(scroll, height=120)
        self._s3_log.grid(row=5, column=0, sticky="ew", pady=(8, 0))

        self._make_footer(frame, self._s3_next)
        self._step_frames[3] = frame

    def _run_s3_clean(self):
        logger = self._logger(self._s3_log)
        self._s3_prog.set(0.05)
        self._s3_lbl.configure(text="Cleaning data …", text_color="gray55")

        def work():
            try:
                df, stats = DataCleaner.clean(
                    self.app_state.raw_df.copy(),
                    self.app_state.column_mapping,
                    logger,
                    progress_cb=lambda p: self.after(0, lambda: self._s3_prog.set(p)),
                    keep_spaces=self.app_state.address_keep_spaces,
                    dob_source_formats=self.app_state.dob_column_format,
                    dob_desired_formats=self.app_state.dob_desired_format,
                )
                self.app_state.cleaned_df    = df
                self.app_state.cleaning_stats = stats
                self.after(0, lambda: self._s3_done(stats, df))
            except Exception as exc:
                logger.error(str(exc))
                import traceback; traceback.print_exc()

        threading.Thread(target=work, daemon=True).start()

    def _set_address_keep_spaces(self, keep_spaces: bool):
        keep_spaces = bool(keep_spaces)
        self.app_state.address_keep_spaces = keep_spaces
        for attr in ("_s3_addr_keep_var", "_s4_addr_keep_var"):
            var = getattr(self, attr, None)
            if var is not None:
                try:
                    if bool(var.get()) != keep_spaces:
                        var.set(keep_spaces)
                except Exception:
                    pass

    def _s3_done(self, stats: dict, df: pd.DataFrame):
        self._s3_prog.set(1.0)
        self._s3_lbl.configure(text="✅  Cleaning complete!", text_color="#4CAF50")
        self._render_stat_cards(self._s3_cards, [
            ("Original Rows",      f"{stats.get('original',0):,}",          "📊", "#6C63FF"),
            ("Valid Rows",         f"{stats.get('final',0):,}",              "✅", "#4CAF50"),
            ("Duplicates Removed", f"{stats.get('duplicates_removed',0):,}", "🗑️", "#FF9800"),
            ("Invalid Removed",    f"{stats.get('invalid_removed',0):,}",    "⚠️", "#F44336"),
        ])
        self._s3_preview.load_dataframe(df)

    def _s3_next(self):
        if self.app_state.cleaned_df is None:
            messagebox.showwarning("Not Ready", "Please wait for cleaning to finish.")
            return
        self._advance()

    # ── STEP 4 ────────────────────────────────────────────────────────────

    def _build_step4(self):
        frame = self._new_step_frame()
        self._make_header(frame, "Step 4 — Settings",
                          "Configure school info, output format, and formula options")

        scroll = ctk.CTkScrollableFrame(frame, fg_color="transparent")
        scroll.grid(row=1, column=0, sticky="nsew", padx=22, pady=8)
        scroll.grid_columnconfigure(0, weight=1)
        scroll.grid_columnconfigure(1, weight=1)

        sc = ctk.CTkFrame(scroll, corner_radius=12)
        sc.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=(0, 10))
        sc.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(sc, text="🏫  School Information",
                     font=ctk.CTkFont(size=13, weight="bold")).grid(
            row=0, column=0, sticky="w", padx=16, pady=(14, 8))

        ctk.CTkLabel(sc, text="School Name *").grid(row=1, column=0, sticky="w", padx=16)
        self._s4_school = ctk.CTkEntry(sc, placeholder_text="e.g. Sunrise Academy",
                                        height=36, corner_radius=8)
        self._s4_school.grid(row=2, column=0, sticky="ew", padx=16, pady=(4, 10))

        ctk.CTkLabel(sc, text="Email Domain").grid(row=3, column=0, sticky="w", padx=16)
        self._s4_domain = ctk.CTkEntry(sc, placeholder_text="school.com",
                        height=36, corner_radius=8)
        self._s4_domain.insert(0, "school.com")
        self._s4_domain.grid(row=4, column=0, sticky="ew", padx=16, pady=(4, 10))

        ctk.CTkLabel(sc, text="Output Format").grid(row=5, column=0, sticky="w", padx=16)
        self._s4_fmt = ctk.CTkComboBox(sc, values=["XLSX", "CSV"], height=34, width=130)
        self._s4_fmt.grid(row=6, column=0, sticky="w", padx=16, pady=(4, 10))

        ctk.CTkLabel(sc, text="Academic Year Formula").grid(row=7, column=0, sticky="w", padx=16)
        self._s4_acad = ctk.CTkEntry(sc, placeholder_text="=1 or =2083 or =YEAR(TODAY())",
                          height=34, corner_radius=8)
        self._s4_acad.insert(0, "=1")
        self._s4_acad.grid(row=8, column=0, sticky="ew", padx=16, pady=(4, 14))

        preset_row = ctk.CTkFrame(sc, fg_color="transparent")
        preset_row.grid(row=9, column=0, sticky="ew", padx=16, pady=(0, 14))
        ctk.CTkButton(preset_row, text="💾  Save Preset", height=30, width=110,
                      corner_radius=7, command=self._s4_save_preset).pack(side="left", padx=(0, 6))
        ctk.CTkButton(preset_row, text="📂  Load Preset", height=30, width=110,
                      corner_radius=7, fg_color="transparent",
                      border_width=1, border_color="#6C63FF",
                      command=self._s4_load_preset).pack(side="left")

        fc = ctk.CTkFrame(scroll, corner_radius=12)
        fc.grid(row=0, column=1, sticky="nsew", padx=(8, 0), pady=(0, 10))
        fc.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(fc, text="🔢  Regd_no Formula",
                     font=ctk.CTkFont(size=13, weight="bold")).grid(
            row=0, column=0, sticky="w", padx=16, pady=(14, 8))

        self._s4_regd_var = tk.StringVar(value="=ROW()-1")
        opts = ["=ROW()-1", '="000"&ROW()-1', "Use Existing Column", "Custom Formula"]
        for i, opt in enumerate(opts, start=1):
            rb = ctk.CTkRadioButton(fc, text=opt, variable=self._s4_regd_var,
                                     value=opt, command=self._s4_regd_changed)
            rb.grid(row=i, column=0, sticky="w", padx=26, pady=3)

        ctk.CTkLabel(fc, text="Custom formula (if selected above):").grid(
            row=5, column=0, sticky="w", padx=16, pady=(8, 2))
        self._s4_custom = ctk.CTkEntry(fc, placeholder_text='=TEXT(ROW()-1,"0000")',
                                        height=32, corner_radius=6, state="disabled")
        self._s4_custom.grid(row=6, column=0, sticky="ew", padx=16, pady=(0, 14))

        # Combo to pick existing source column for Regd_no when selected
        self._s4_regd_col_var = tk.StringVar(value=self.app_state.regd_no_column)
        self._s4_regd_col_cb = ctk.CTkComboBox(fc, values=[], variable=self._s4_regd_col_var,
                               width=200, height=28, state="disabled")
        self._s4_regd_col_cb.grid(row=7, column=0, sticky="w", padx=26, pady=(0, 12))

        pv = ctk.CTkFrame(scroll, corner_radius=12)
        pv.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 10))
        pv.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(pv, text="👁️  Formula Preview",
                     font=ctk.CTkFont(size=13, weight="bold")).grid(
            row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(14, 6))

        self._s4_preview_box = ctk.CTkTextbox(
            pv, height=90, corner_radius=6, state="disabled",
            font=ctk.CTkFont(family="Consolas", size=11),
        )
        self._s4_preview_box.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 4))

        ctk.CTkButton(pv, text="🔄  Refresh Preview", height=30, width=150,
                      corner_radius=7, command=self._s4_refresh).grid(
            row=2, column=0, sticky="w", padx=16, pady=(0, 14))

        # Presets panel (hidden by default) shown inline near formula area
        self._s4_presets_panel = ctk.CTkScrollableFrame(fc, height=240, corner_radius=8)
        self._s4_presets_panel.grid_columnconfigure(0, weight=1)
        self._s4_presets_panel.grid(row=8, column=0, sticky="ew", padx=16, pady=(6, 12))
        self._s4_presets_panel.grid_remove()

        self._make_footer(frame, self._s4_next)
        self._step_frames[4] = frame

    def _s4_regd_changed(self):
        val = self._s4_regd_var.get()
        self._s4_custom.configure(state="normal" if val == "Custom Formula" else "disabled")
        # enable/disable existing-column combobox
        try:
            if hasattr(self, "_s4_regd_col_cb") and self._s4_regd_col_cb is not None:
                state = "normal" if val == "Use Existing Column" else "disabled"
                self._s4_regd_col_cb.configure(state=state)
        except Exception:
            pass

    def _s4_refresh(self):
        dom  = self._s4_domain.get() or "school.com"
        acad = self._s4_acad.get() or "=1"
        regd = self._s4_regd_var.get()
        text = (
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"Regd_no       : {regd}\n"
            f"Email         : first name + class token + row suffix @ {dom}\n"
            f"Academic Year : {acad.lstrip('=')}\n"
            f"Status        : 1  (always)\n"
            f"roll_number   : =ROW()-1\n"
            f"shift         : 2  (always)\n"
            f"DOB format    : YYYYMMDD (separators removed)\n"
            f"Gender        : Male→1  Female→2  Others→3\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
        )
        self._s4_preview_box.configure(state="normal")
        self._s4_preview_box.delete("1.0", "end")
        self._s4_preview_box.insert("1.0", text)
        self._s4_preview_box.configure(state="disabled")

    def _s4_save_preset(self):
        name = self._s4_school.get().strip() or "default"
        self._sync_settings()
        self.app_state.save_preset(name)
        messagebox.showinfo("Preset Saved", f"Preset '{name}' saved.")

    def _s4_load_preset(self):
        # Toggle inline presets panel
        presets = self.app_state.list_presets()
        if not presets:
            messagebox.showinfo("No Presets", "No saved presets found.")
            return
        panel = getattr(self, "_s4_presets_panel", None)
        if panel is None:
            messagebox.showerror("Error", "Presets panel not available.")
            return
        # if already visible, hide it
        if panel.winfo_ismapped():
            panel.grid_remove()
            return

        # clear and populate
        for w in panel.winfo_children():
            w.destroy()

        ctk.CTkLabel(panel, text="Saved Presets", font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, column=0, sticky="w", pady=(8,6), padx=8)

        # limit list to current presets (AppState.save_preset enforces FIFO on save)
        for i, name in enumerate(presets):
            if i >= 7:
                break
            rowf = ctk.CTkFrame(panel, fg_color="transparent")
            rowf.grid(row=i+1, column=0, sticky="ew", padx=8, pady=4)
            rowf.grid_columnconfigure(0, weight=1)
            lbl = ctk.CTkLabel(rowf, text=name, anchor="w")
            lbl.grid(row=0, column=0, sticky="w")

            def _load(n=name):
                if self.app_state.load_preset(n):
                    self._s4_school.delete(0, "end"); self._s4_school.insert(0, self.app_state.school_name)
                    self._s4_domain.delete(0, "end"); self._s4_domain.insert(0, self.app_state.school_domain)
                    self._s4_acad.delete(0, "end");   self._s4_acad.insert(0, self.app_state.academic_year)
                    self._s4_fmt.set(self.app_state.output_format)
                    self._s4_regd_var.set(self.app_state.regd_no_mode)
                    panel.grid_remove()

            def _delete(n=name, rf=rowf):
                if messagebox.askyesno("Delete", f"Delete preset '{n}'?"):
                    if self.app_state.delete_preset(n):
                        rf.destroy()

            def _start_rename(n=name, lbl_widget=lbl, rf=rowf):
                # replace label with entry + save/cancel
                entry = ctk.CTkEntry(rf, width=180)
                entry.insert(0, n)
                entry.grid(row=0, column=0, sticky="w")
                btn_save = ctk.CTkButton(rf, text="Save", width=60, command=lambda: _rename_confirm(n, entry.get(), rf))
                btn_save.grid(row=0, column=1, padx=6)
                btn_cancel = ctk.CTkButton(rf, text="Cancel", width=60, command=lambda: _rename_cancel(n, lbl_widget, entry, btn_save, btn_cancel))
                btn_cancel.grid(row=0, column=2, padx=6)

            def _rename_confirm(old, new, rf=rowf):
                new = new.strip()
                if not new or new == old:
                    messagebox.showinfo("Rename", "No change.")
                    return
                if self.app_state.rename_preset(old, new):
                    # refresh panel
                    self._s4_load_preset()
                else:
                    messagebox.showerror("Rename", "Rename failed (name exists or error).")

            def _rename_cancel(old, lbl_widget, entry, btn_save, btn_cancel):
                entry.destroy(); btn_save.destroy(); btn_cancel.destroy(); lbl_widget.grid(row=0, column=0, sticky="w")

            ctk.CTkButton(rowf, text="Load", width=60, command=_load).grid(row=0, column=1, padx=6)
            ctk.CTkButton(rowf, text="Rename", width=60, command=_start_rename).grid(row=0, column=2, padx=6)
            ctk.CTkButton(rowf, text="Delete", width=60, command=_delete).grid(row=0, column=3, padx=6)

        panel.grid()

    def _sync_settings(self):
        self.app_state.school_name   = self._s4_school.get().strip()
        self.app_state.school_domain = self._s4_domain.get().strip() or "school.com"
        self.app_state.output_format = self._s4_fmt.get()
        self.app_state.regd_no_mode  = self._s4_regd_var.get()
        self.app_state.regd_no_custom = self._s4_custom.get().strip()
        self.app_state.academic_year = self._s4_acad.get().strip() or "=1"
        try:
            self.app_state.regd_no_column = self._s4_regd_col_var.get().strip()
        except Exception:
            self.app_state.regd_no_column = ""

    def _s4_next(self):
        if not self._s4_school.get().strip():
            messagebox.showwarning("Required", "Please enter the school name.")
            return
        self._sync_settings()
        self._advance()
        self._run_s5_erp()

    # ── STEP 5 ────────────────────────────────────────────────────────────

    def _build_step5(self):
        frame = self._new_step_frame()
        self._make_header(frame, "Step 5 — ERP Reformat Engine",
                          "Build the 19-column ERP structure with auto-generated formulas")

        scroll = ctk.CTkScrollableFrame(frame, fg_color="transparent")
        scroll.grid(row=1, column=0, sticky="nsew", padx=22, pady=8)
        scroll.grid_columnconfigure(0, weight=1)

        self._s5_cards = ctk.CTkFrame(scroll, fg_color="transparent")
        self._s5_cards.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        self._s5_prog = ctk.CTkProgressBar(scroll, height=8, corner_radius=4)
        self._s5_prog.grid(row=1, column=0, sticky="ew", pady=(0, 4))
        self._s5_prog.set(0)

        self._s5_lbl = ctk.CTkLabel(scroll, text="Building ERP …", text_color="gray55",
                                     font=ctk.CTkFont(size=11))
        self._s5_lbl.grid(row=2, column=0, sticky="w", pady=(0, 10))

        self._s5_val_box = ctk.CTkTextbox(scroll, height=70, corner_radius=8,
                                           state="disabled", font=ctk.CTkFont(size=11))
        self._s5_val_box.grid(row=3, column=0, sticky="ew", pady=(0, 8))

        ctk.CTkLabel(scroll, text="ERP Structure Preview (19 Columns)",
                     font=ctk.CTkFont(size=13, weight="bold")).grid(
            row=4, column=0, sticky="w", pady=(0, 4))

        self._s5_preview = PreviewTable(scroll)
        self._s5_preview.grid(row=5, column=0, sticky="nsew")
        scroll.grid_rowconfigure(5, weight=1)

        self._s5_log = LogPanel(scroll, height=120)
        self._s5_log.grid(row=6, column=0, sticky="ew", pady=(8, 0))

        self._make_footer(frame, self._s5_next)
        self._step_frames[5] = frame

    def _run_s5_erp(self):
        logger = self._logger(self._s5_log)
        self._s5_prog.set(0.05)
        self._s5_lbl.configure(text="Building ERP structure …", text_color="gray55")

        def work():
            try:
                df = self.app_state.cleaned_df.copy()
                df.columns = [
                    c.replace(" ", "_").replace("'", "").replace("-", "_")
                    for c in df.columns
                ]
                erp = FormulaEngine.build_erp(
                    df, self.app_state, logger,
                    progress_cb=lambda p: self.after(0, lambda: self._s5_prog.set(p)),
                )
                issues = ValidationEngine.validate(erp)
                self.app_state.erp_df = erp
                self.after(0, lambda: self._s5_done(erp, issues))
            except Exception as exc:
                logger.error(str(exc))
                import traceback; traceback.print_exc()

        threading.Thread(target=work, daemon=True).start()

    def _s5_done(self, df: pd.DataFrame, issues: List[str]):
        self._s5_prog.set(1.0)
        self._s5_lbl.configure(text="✅  ERP structure ready!", text_color="#4CAF50")

        n_classes = 0
        if self.app_state.cleaned_df is not None and "CurrentClass" in self.app_state.cleaned_df.columns:
            n_classes = self.app_state.cleaned_df["CurrentClass"].nunique()

        self._render_stat_cards(self._s5_cards, [
            ("ERP Records", f"{len(df):,}",       "📊", "#6C63FF"),
            ("ERP Columns", str(len(df.columns)), "📋", "#4CAF50"),
            ("Classes",     str(n_classes),       "📚", "#FF9800"),
            ("Formulas",    "6",                  "🔢", "#2196F3"),
        ])

        val_text = "\n".join(issues)
        self._s5_val_box.configure(state="normal")
        self._s5_val_box.delete("1.0", "end")
        self._s5_val_box.insert("1.0", val_text)
        self._s5_val_box.configure(state="disabled")

        self._s5_preview.load_dataframe(df)

    def _s5_next(self):
        if self.app_state.erp_df is None:
            messagebox.showwarning("Not Ready", "Please wait for ERP formatting.")
            return
        self._advance()
        self._run_s6_separate()

    # ── STEP 6 ────────────────────────────────────────────────────────────

    def _build_step6(self):
        frame = self._new_step_frame()
        self._make_header(frame, "Step 6 — Class-Wise Separation",
                          "Separate students by class and sort alphabetically")

        scroll = ctk.CTkScrollableFrame(frame, fg_color="transparent")
        scroll.grid(row=1, column=0, sticky="nsew", padx=22, pady=8)
        scroll.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(scroll, text="📚  Detected Classes",
                     font=ctk.CTkFont(size=13, weight="bold")).grid(
            row=0, column=0, sticky="w", pady=(0, 4))

        self._s6_class_frame = ctk.CTkFrame(scroll, corner_radius=12)
        self._s6_class_frame.grid(row=1, column=0, sticky="ew", pady=(0, 12))
        self._s6_class_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(scroll, text="Class Preview",
                     font=ctk.CTkFont(size=13, weight="bold")).grid(
            row=2, column=0, sticky="w", pady=(0, 4))

        self._s6_preview = PreviewTable(scroll)
        self._s6_preview.grid(row=3, column=0, sticky="nsew")
        scroll.grid_rowconfigure(3, weight=1)

        self._s6_log = LogPanel(scroll, height=100)
        self._s6_log.grid(row=4, column=0, sticky="ew", pady=(8, 0))

        self._make_footer(frame, self._s6_next, label="Choose Output Folder →")
        self._step_frames[6] = frame

    def _run_s6_separate(self):
        logger = self._logger(self._s6_log)

        def work():
            try:
                df_clean = self.app_state.cleaned_df
                df_erp   = self.app_state.erp_df.copy()
                # Determine class and section columns (support both legacy and new names)
                if df_clean is None or all(c not in df_clean.columns for c in ("CurrentClass", "Class")):
                    logger.warn("No class column found — treating all as one group")
                    df_erp = df_erp.reset_index(drop=True)
                    df_erp["roll_number"] = range(1, len(df_erp) + 1)
                    self.app_state.class_groups = {"All_Students": df_erp}
                    self.after(0, lambda: self._s6_done({"All_Students": df_erp}))
                    return

                if "Class" in df_clean.columns:
                    classes = df_clean["Class"].reset_index(drop=True)
                else:
                    classes = df_clean["CurrentClass"].reset_index(drop=True)
                if len(classes) > len(df_erp):
                    classes = classes.iloc[:len(df_erp)]
                elif len(classes) < len(df_erp):
                    extra = pd.Series(["Unknown"] * (len(df_erp) - len(classes)))
                    classes = pd.concat([classes, extra], ignore_index=True)

                # Section handling
                if df_clean is not None and "Section" in df_clean.columns:
                    sections = df_clean["Section"].reset_index(drop=True)
                    if len(sections) > len(df_erp):
                        sections = sections.iloc[:len(df_erp)]
                    elif len(sections) < len(df_erp):
                        extra = pd.Series([""] * (len(df_erp) - len(sections)))
                        sections = pd.concat([sections, extra], ignore_index=True)
                else:
                    sections = pd.Series([""] * len(df_erp))

                df_erp["_cls"] = classes.fillna("Unknown").astype(str).values
                df_erp["_sec"] = sections.fillna("").astype(str).values
                groups: Dict[str, pd.DataFrame] = {}

                for (cls, sec), grp in df_erp.groupby(["_cls", "_sec"]):
                    cls_key = str(cls).strip() or "Unknown"
                    sec_key = str(sec).strip() or "default"
                    grp = grp.drop(columns=["_cls", "_sec"]).sort_values("Name").reset_index(drop=True)
                    grp["roll_number"] = range(1, len(grp) + 1)
                    group_name = f"{cls_key} - {sec_key}" if sec_key and sec_key != "default" else cls_key
                    # Normalize filename safe keys later in export
                    groups[group_name] = grp
                    logger.success(f"Group '{group_name}': {len(grp):,} students")

                def _num(k: str) -> int:
                    m = re.search(r"\d+", k)
                    return int(m.group()) if m else 9999

                groups = dict(sorted(groups.items(), key=lambda x: _num(x[0])))
                self.app_state.class_groups = groups
                self.after(0, lambda: self._s6_done(groups))

            except Exception as exc:
                logger.error(str(exc))
                import traceback; traceback.print_exc()

        threading.Thread(target=work, daemon=True).start()

    def _s6_done(self, groups: Dict[str, pd.DataFrame]):
        for w in self._s6_class_frame.winfo_children():
            w.destroy()

        school = self.app_state.school_name or "School"
        ext    = self.app_state.output_format.lower()

        for i, (cls, df) in enumerate(groups.items()):
            cls_safe = ExportEngine._class_filename_token(cls)
            sch_safe = re.sub(r"[^\w]", "_", school)
            fname    = f"{sch_safe}_{cls_safe}.{ext}"

            row = ctk.CTkFrame(self._s6_class_frame, fg_color="transparent")
            row.grid(row=i, column=0, sticky="ew", padx=12, pady=4)
            row.grid_columnconfigure(2, weight=1)

            ctk.CTkLabel(row, text=f"📚  {cls}", width=150,
                         font=ctk.CTkFont(size=12, weight="bold"),
                         anchor="w").grid(row=0, column=0, padx=(0, 10))
            ctk.CTkLabel(row, text=f"{len(df):,} students",
                         text_color="#4CAF50", width=100).grid(row=0, column=1)
            ctk.CTkLabel(row, text=fname, text_color="gray55",
                         font=ctk.CTkFont(size=10)).grid(row=0, column=2, sticky="w")
            ctk.CTkButton(
                row, text="Preview", width=72, height=26, corner_radius=6,
                fg_color="transparent", border_width=1, border_color="#6C63FF",
                command=lambda d=df: self._s6_preview.load_dataframe(d),
            ).grid(row=0, column=3, padx=(10, 0))

        if groups:
            self._s6_preview.load_dataframe(next(iter(groups.values())))

    def _s6_next(self):
        if not self.app_state.class_groups:
            messagebox.showwarning("Not Ready", "Please wait for class separation.")
            return
        folder = filedialog.askdirectory(title="Select Output Folder for Exported Files")
        if not folder:
            return
        self.app_state.output_folder = folder
        self._advance()
        self._run_s7_export()

    # ── STEP 7 ────────────────────────────────────────────────────────────

    def _build_step7(self):
        frame = self._new_step_frame()
        self._make_header(frame, "Step 7 — Export Engine",
                          "Write class-wise files with styled headers, formulas, and filters")

        scroll = ctk.CTkScrollableFrame(frame, fg_color="transparent")
        scroll.grid(row=1, column=0, sticky="nsew", padx=22, pady=8)
        scroll.grid_columnconfigure(0, weight=1)

        self._s7_prog = ctk.CTkProgressBar(scroll, height=12, corner_radius=6)
        self._s7_prog.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        self._s7_prog.set(0)

        self._s7_lbl = ctk.CTkLabel(scroll, text="Preparing …",
                                     text_color="gray55",
                                     font=ctk.CTkFont(size=13))
        self._s7_lbl.grid(row=1, column=0, sticky="w", pady=(0, 12))

        self._s7_cards = ctk.CTkFrame(scroll, fg_color="transparent")
        self._s7_cards.grid(row=2, column=0, sticky="ew", pady=(0, 12))

        ctk.CTkLabel(scroll, text="📁  Exported Files",
                     font=ctk.CTkFont(size=13, weight="bold")).grid(
            row=3, column=0, sticky="w", pady=(0, 4))

        self._s7_files = ctk.CTkFrame(scroll, corner_radius=12)
        self._s7_files.grid(row=4, column=0, sticky="ew", pady=(0, 12))
        self._s7_files.grid_columnconfigure(0, weight=1)

        self._s7_log = LogPanel(scroll, height=140)
        self._s7_log.grid(row=5, column=0, sticky="ew", pady=(0, 10))

        self._s7_open_btn = ctk.CTkButton(
            scroll, text="📂  Open Output Folder",
            height=46, corner_radius=12,
            fg_color="#4CAF50", hover_color="#388E3C",
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._s7_open_folder,
        )
        self._s7_open_btn.grid(row=6, column=0, sticky="ew", pady=(0, 8))

        ctk.CTkButton(
            scroll, text="🔄  Start New Session",
            height=38, corner_radius=10,
            fg_color="transparent", border_width=1, border_color="#6C63FF",
            command=self._restart,
        ).grid(row=7, column=0, sticky="ew")

        frame.grid_rowconfigure(2, weight=0)
        self._step_frames[7] = frame

    def _run_s7_export(self):
        logger = self._logger(self._s7_log)
        self._s7_prog.set(0.04)
        self._s7_lbl.configure(text="Exporting files …", text_color="gray55")

        def work():
            try:
                # exported → List[Tuple[filepath, original_class_key]]
                exported: List[Tuple[str, str]] = ExportEngine.export_all(
                    self.app_state.class_groups,
                    self.app_state,
                    logger,
                    progress_cb=lambda p: self.after(0, lambda: self._s7_prog.set(p)),
                )
                self.after(0, lambda: self._s7_done(exported))
            except Exception as exc:
                logger.error(str(exc))
                import traceback; traceback.print_exc()

        threading.Thread(target=work, daemon=True).start()

    def _s7_done(self, exported: List[Tuple[str, str]]):
        # exported is a list of (filepath, original_class_key) tuples
        self._s7_prog.set(1.0)
        total = sum(len(d) for d in self.app_state.class_groups.values())
        self._s7_lbl.configure(
            text=f"✅  Export complete!  {len(exported)} files  ·  {total:,} total records",
            text_color="#4CAF50",
        )
        self._render_stat_cards(self._s7_cards, [
            ("Files Exported",  str(len(exported)),                    "💾", "#4CAF50"),
            ("Total Students",  f"{total:,}",                         "👨‍🎓", "#6C63FF"),
            ("Classes",         str(len(self.app_state.class_groups)), "📚", "#FF9800"),
            ("Format",          self.app_state.output_format,         "📄", "#2196F3"),
        ])

        for w in self._s7_files.winfo_children():
            w.destroy()
        for i, (fp, cls_key) in enumerate(exported):
            n_rows = len(self.app_state.class_groups.get(cls_key, pd.DataFrame()))
            r = ctk.CTkFrame(self._s7_files, fg_color="transparent")
            r.grid(row=i, column=0, sticky="ew", padx=12, pady=3)
            r.grid_columnconfigure(1, weight=1)
            ctk.CTkLabel(r, text="✅", width=26).grid(row=0, column=0)
            ctk.CTkLabel(r, text=os.path.basename(fp),
                         font=ctk.CTkFont(size=11), anchor="w").grid(
                row=0, column=1, sticky="w", padx=6)
            ctk.CTkLabel(r, text=f"{n_rows:,} rows",
                         text_color="gray55", font=ctk.CTkFont(size=10)).grid(
                row=0, column=2, padx=(0, 10))

    def _s7_open_folder(self):
        folder = self.app_state.output_folder
        if folder and os.path.isdir(folder):
            if sys.platform == "win32":
                os.startfile(folder)
            elif sys.platform == "darwin":
                subprocess.run(["open", folder])
            else:
                subprocess.run(["xdg-open", folder])
        else:
            messagebox.showinfo("Folder", "Output folder not accessible.")

    def _restart(self):
        if messagebox.askyesno("Start Over?",
                               "This will clear all current data. Continue?"):
            self.app_state = AppState()
            self._show_step(1)
            self._s1_prog.set(0)
            self._s1_prog_lbl.configure(text="", text_color="gray55")

    def _new_step_frame(self) -> ctk.CTkFrame:
        f = ctk.CTkFrame(self._main, corner_radius=0, fg_color="transparent")
        f.grid_columnconfigure(0, weight=1)
        f.grid_rowconfigure(1, weight=1)
        return f

    def _on_close(self):
        self.destroy()


def _divider(parent):
    ctk.CTkFrame(parent, height=1, fg_color=("gray82", "#252843")).pack(
        fill="x", padx=14, pady=6)


def _strip_ext(fname: str, school: str) -> str:
    base = os.path.splitext(fname)[0]
    prefix = re.sub(r"[^\w]", "_", school or "School") + "_"
    if base.startswith(prefix):
        return base[len(prefix):].replace("_", " ")
    return base


def main():
    PRESET_DIR.mkdir(parents=True, exist_ok=True)
    app = SchoolERPApp()
    app.mainloop()


if __name__ == "__main__":
    main()